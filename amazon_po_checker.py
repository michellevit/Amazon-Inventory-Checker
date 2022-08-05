import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from collections import OrderedDict
import os
from datetime import date
from string import ascii_lowercase

#sanity check

# Check if AMAZON US .xlsx file exists
## If true, open first sheet
## If false, print error message and name it
if os.path.exists("EditLineItems.xlsx"):
    wb_us = load_workbook(filename = 'EditLineItems.xlsx')
    sheet_us = wb_us.active
else: 
    print("Error - Filename: 'EditLineItems.xlsx' does not exist.")
    sheet_us = "Does not exist"


# Check if AMAZON CA .xlsx file exists
## If true, open first sheet
## If false, name it
if os.path.exists("EditLineItems (1).xlsx"):
    wb_ca = load_workbook(filename = 'EditLineItems (1).xlsx')
    sheet_ca = wb_ca.active
else: 
    print("Error - Filename: 'EditLineItems (1).xlsx' does not exist.")
    sheet_ca = "Does not exist"


# Test to make sure the file is compatible with this program, if it exists
def correct_wb(sheet):
    if (sheet == "Does not exist"):
        return
    else:
        assert sheet['A1'].value == 'PO',  'Cell A1 should be Model Number'
        assert sheet['G1'].value == 'Model Number',  'Cell G1 should be Model Number'
        assert sheet['O1'].value == 'Expected Quantity', 'Cell O1 should be Expected Quantity'
        assert sheet['P1'].value == 'Unit Cost', 'Cell P1 should be Unit Cost'
correct_wb(sheet_ca)
correct_wb(sheet_us)


# Make a list of all the Canadian PO Numbers, so you can keep track of them later
cad_pos = []
if (sheet_ca != "Does not exist"):
    for row in sheet_ca.iter_rows(min_row=2, max_row=(sheet_ca.max_row)):
        po = row[0].value
        if po not in cad_pos:
            cad_pos.append(po)


# Create a new excel excel workbook/file (at the end), containing the combined data from the US + CA orders (if they exist)
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = 'POs over $400'
today = date.today()
new_filename = "POs to Confirm " + today.strftime("%B %d, %Y") + ".xlsx"
new_wb.save(filename = new_filename)
## Copy raw data from sheet_us to new sheet
for i in range(1, sheet_us.max_row+1):
    for j in range(1, sheet_us.max_column+1):
        new_sheet.cell(row=i, column=j).value = sheet_us.cell(row=i, column=j).value
## Append raw data from sheet_ca to new sheet
if (sheet_ca != "Does not exist"):
    new_sheet_end = new_sheet.max_row - 1
    for i in range(2, sheet_ca.max_row +1):
        for j in range(1, sheet_ca.max_column + 1):
            new_sheet.cell(row=i+new_sheet_end, column=j).value = sheet_ca.cell(row=i, column=j).value
## Save file
new_wb.save(new_filename)


# Create a dictionary to keep track of the total value of each PO (key=PO, value=total PO value)
po_dict = {}
# Calculate each POs total value, by iterating through values in column A and U (i.e. PO and Unit Cost), and save the data to po_dict
for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row):
    po = row[0].value
    cost = round(row[14].value * row[15].value) # cost times quantity
    ## If po in dictionary, update quantity
    if po in po_dict.keys():
        po_dict[po] = round(po_dict[po] + cost)
    ## If po not in dictionary, add it to dictionary
    else:
        po_dict[po] = cost


# Sort and format the po_dict data, then print it to the console

## Sort POs by value, in descending order
sorted_po_dict = sorted(po_dict.items(), key=lambda x: x[1], reverse=True)

## Make a new list of orders to be cancelled
orders_to_cancel = []

## Define the minimum threshold of order cost
min_po_value = 380

## Print the sorted dictionary
print('---------------\nPOs to Confirm/Cancel\n---------------')
for value in sorted_po_dict:
    ## If the order is from Canada, format it with '- CA' at the end, or else add blank spaces to the end
    if (value[0] in cad_pos):
        po_formatted = value[0] + ' (CAD) : '
    else: 
        po_formatted = value[0] + '       : '
    ## If the order is above 380, print it out with no additional note
    if value[1] >= min_po_value:
        print(po_formatted, value[1])
    ## If order value is below the min_po_value, format it with a note to cancel it
    else: 
        cost_formatted = str(value[1])
        spaces_needed = 5 - len(cost_formatted)
        for space in range(spaces_needed):
            cost_formatted = cost_formatted + ' '
        print(po_formatted, cost_formatted, '- CANCEL')
        ## Add the order number to a list of orders to be cancelled
        orders_to_cancel.append(value[0])        


# Delete the rows in new_wb which contain POs that are in the orders_to_cancel list, starting from the bottom up
rows = list(new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row))
rows = reversed(rows)
for row in rows:
    if row[0].value in orders_to_cancel:
        new_sheet.delete_rows(row[0].row, 1)
## Save file
new_wb.save(new_filename)


# Create a dictionary to keep track of the unit quantity ordered for each item (key=model number, value=units ordered)
## Create empty dictionary
inventory_dict = {}
## Iterate through values in column G and N (i.e. Model Number and Expected Quantity)
for row in new_sheet.iter_rows(min_row=2, max_row=(new_sheet.max_row)):
    product = row[6].value
    # Format the product name, so that it appears cleaner when printed 
    spaces_needed = 6 - len(product)
    for space in range(spaces_needed):
        product = product + ' '
    quantity = round((row[14].value))
    # If product in dictionary, update quantity
    if product in inventory_dict.keys():
        inventory_dict[product] = inventory_dict[product] + quantity
    # If product not in dictionary, add it to dictionary
    else:
        inventory_dict[product] = quantity


# Format the invevenory dictionary
## Sort the inventory dictionary alphabetically
sorted_inventory_dict = OrderedDict(sorted(inventory_dict.items(), key=lambda t: t[0]))
## Print the sorted dictionary
print('---------------\nInventory Requested (from non-cancelled POs)\n---------------')
for key, value in sorted_inventory_dict.items():
        print(key, ':', value)


# Create a new sheet (in new_wb), with the POs and Requested Units
inventory_sheet = new_wb.create_sheet('Inventory to Confirm')
new_wb.active = new_wb['Inventory to Confirm']
inventory_sheet.cell(row=1, column=1).value = "Model Number"
inventory_sheet.cell(row=1, column=2).value = "Requested Units"
inventory_sheet.cell(row=1, column=3).value = "Quantity Accepted"


for row in inventory_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")
        x = cell.coordinate
        inventory_sheet[x].font = Font(bold=True)
        inventory_sheet.column_dimensions[x[0]].width = 20
        if x == 'C1':
            cell.fill = PatternFill("solid", start_color="FFFF00")
        else: 
            cell.fill = PatternFill("solid", start_color="ADD8E6")
        


## Save file
new_wb.save(new_filename)



# Make POs to confirm pretty -> highlight POs and make column A wider
# Add note in 'inventory to confirm' sheet saying that this is inventory from orders over $400 / minimum threshold
# Fill in 'Inventory to Confirm' sheet, with data
# new .py file which will 
## compare the quantity rejected with requested units
### then either remove them all from POs and recalculate POs over $400, and output data into new sheet 'POs to Confirm'
### or it will evaluate which POs to take the inventory from, and output data into new sheet 'POs to Confirm'
###














# Delete all save files except one
