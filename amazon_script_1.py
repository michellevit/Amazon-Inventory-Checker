import os, sys
from os import path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment
from datetime import date
import operator


# GLOBAL VARIABLE
min_po_value = 100


def pos_to_confirm():
    global min_po_value
    # Get Amazon Vendor Download file
    amazon_files = get_amazon_files()
    # Open workbooks and make sure they comply with this script's requirements
    file_1 = amazon_files[0]
    wb_1 = load_workbook(filename=file_1)
    ws_1 = wb_1.active
    check_wb(ws_1)
    amazon_files[0] = ws_1
    if len(amazon_files) == 2:
        file_2 = amazon_files[1]
        wb_2 = load_workbook(filename=file_2)
        ws_2 = wb_2.active
        check_wb(ws_2)
        amazon_files[1] = ws_2
    # Create new workbook
    new_wb_data = create_new_wb()
    raw_data_sheet = new_wb_data[0]
    new_wb = new_wb_data[1]
    # Copy data from original Amazon files into new workbook
    raw_data_sheet = populate_new_wb(raw_data_sheet, amazon_files)
    # Create list of POs with their total PO value
    sorted_po_list = get_and_sort_po_values(raw_data_sheet)
    # Create list of POs under min po value
    orders_to_cancel_list = create_list_pos_to_cancel(sorted_po_list)
    # Create list of CAD POs
    cad_po_list = get_cad_pos(raw_data_sheet)
    # Print PO list to the console
    po_heading_title = (
        "---------------\nAmazon: POs to Confirm/Cancel (unconfirmed)\n---------------"
    )
    # print_pos_to_confirm(
    #     sorted_po_list, cad_po_list, orders_to_cancel_list, po_heading_title
    # )
    # Create inventory dicts to track inventory in stock (2nd sheet)
    inventory_dicts = create_inventory_tracker_dicts(
        raw_data_sheet, orders_to_cancel_list
    )
    # Print the inventory requested to the console
    inv_heading_title = "---------------\nAmazon: Inventory Requested (from non-cancelled POs - unconfirmed)\n---------------"
    # print_inventory_to_confirm(inventory_dicts, inv_heading_title)
    # Create a new sheet for the 'Inv to Confirm' and populate it
    inventory_sheet = create_inventory_sheet(new_wb)
    populate_inventory_sheet(inventory_sheet, inventory_dicts)
    # Create a sheet for the 'POs to Confirm (unadjusted)' and populate it
    unadjusted_pos_sheet = create_pos_to_confirm_unadjusted_sheet(new_wb)
    populate_pos_to_confirm_unadjusted_sheet(
        raw_data_sheet, unadjusted_pos_sheet, orders_to_cancel_list
    )
    update_raw_data_sheet(raw_data_sheet, orders_to_cancel_list)
    today = date.today()
    new_wb_filename = "POs to Confirm - " + today.strftime("%B %d, %Y") + ".xlsx"
    save_new_wb(new_wb, new_wb_filename)


# Get Amazon files and make sure they comply with this script's requirements
def get_amazon_files():
    my_dir = os.listdir()
    amazon_files = []
    for x in my_dir:
        if "VendorDownload" in x:
            amazon_file = x
            if amazon_file[-4:] != "xlsx":
                raise Exception(
                    "Error: File must be in .xlsx format - instructions to solve:\n 1. Open the .xls file\n 2. Click: File > Save as \n 3. Select the Save as type: .xlsx\n 4. Delete the .xls file"
                )
            else:
                amazon_files.append(amazon_file)
    if len(amazon_files) > 2:
        raise Exception(
            "Error: There should only be 1-2 Amazon Vendor Download file(s) in the directory"
        )
    elif len(amazon_files) == 0:
        raise Exception(
            "Error: There is no Amazon Vendor Donwload file in the directory"
        )
    return amazon_files


# Test to make sure the file is compatible with this program, if it exists
def check_wb(ws):
    print("TEST: ", ws)
    assert ws["A3"].value == "Order/PO Number", "Cell A3 should be Order/PO Number, or you have saved the file as type 'Strict Open XML Spreadhseet (*.xlsx)' instead of as type 'Excel Workbook (*.xlsx)'"
    assert ws["B3"].value == "External ID", "Cell A3 should be External ID"
    assert ws["C3"].value == "Model Number", "Cell C3 should be Model Number"
    assert ws["D3"].value == "ASIN", "Cell D3 should be ASIN"


# Create a new excel excel workbook/file
def create_new_wb():
    new_wb = openpyxl.Workbook()
    raw_data_sheet = new_wb.active
    raw_data_sheet.title = "PO Raw Data"
    return raw_data_sheet, new_wb


# Populate new workbook with data from original Amazon workbooks
def populate_new_wb(raw_data_sheet, amazon_files):
    ws_1 = amazon_files[0]
    for i in range(3, ws_1.max_row + 1):
        for j in range(1, ws_1.max_column + 1):
            raw_data_sheet.cell(row=i - 2, column=j).value = ws_1.cell(
                row=i, column=j
            ).value
    if len(amazon_files) == 2:
        ws_2 = amazon_files[1]
        po_sheet_end = raw_data_sheet.max_row + 1
        k = 0
        for i in range(4, ws_2.max_row + 1):
            for j in range(1, ws_2.max_column + 1):
                raw_data_sheet.cell(row=k + po_sheet_end, column=j).value = ws_2.cell(
                    row=i, column=j
                ).value
            k = k + 1
    return raw_data_sheet


# Create a dictionary with total value of each PO (key=PO, value=total PO value)
def get_and_sort_po_values(raw_data_sheet):
    po_dict = {}
    # Calculate each POs total value, by iterating through values in column A and U (i.e. PO and Unit Cost), and save the data to po_dict
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row):
        if row[0].value == None:
            break
        po = row[0].value
        cost = row[8].value
        quantity_confirmed = row[11].value
        po_value = cost * quantity_confirmed
        # If po in dictionary, update quantity
        if po in po_dict.keys():
            po_dict[po] = round(po_dict[po] + po_value)
        # If po not in dictionary, add it to dictionary
        else:
            po_dict[po] = po_value
    # Sort POs by value, in descending order
    sorted_po_list = sorted(po_dict.items(), key=lambda x: x[1], reverse=True)
    return sorted_po_list


# Create list of POs to cancel
def create_list_pos_to_cancel(sorted_po_list):
    orders_to_cancel_list = []
    for key, value in sorted_po_list:
        if (value < min_po_value) and (key not in orders_to_cancel_list):
            orders_to_cancel_list.append(key)
    return orders_to_cancel_list


def get_cad_pos(raw_data_sheet):
    cad_po_list = []
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row):
        currency = row[31].value
        if currency == "CAD":
            po = row[0].value
            if po not in cad_po_list:
                cad_po_list.append(po)
    return cad_po_list


# Print the sorted 'POs to Confirm/Cancel' dictionary to Console
def print_pos_to_confirm(
    sorted_po_list, cad_po_list, orders_to_cancel_list, po_heading_title
):
    print(po_heading_title)
    longest_po_name = 0
    for item in sorted_po_list:
        po_number = item[0]
        if po_number in cad_po_list:
            po_number = po_number + " (CAD)"
        if len(po_number) > longest_po_name:
            longest_po_name = len(po_number)
    for item in sorted_po_list:
        po_number = item[0]
        po_value = round(item[1])
        spaces_needed = longest_po_name - len(po_number)
        extra_space = ""
        for space in range(spaces_needed):
            extra_space = extra_space + " "
        po_formatted = po_number + extra_space + " : "
        if po_number in orders_to_cancel_list:
            po_formatted = po_number + extra_space + " :"
            if po_number in cad_po_list:
                po_number = po_number + " (CAD)"
            print(po_formatted, "*CANCEL*")
        # If order value is below the min_po_value, format it with a note to cancel it
        else:
            if po_number in cad_po_list:
                po_number = po_number + " (CAD)"
            print(po_formatted, po_value, sep="")


# Keep track of the unit quantity ordered for each item
def create_inventory_tracker_dicts(raw_data_sheet, orders_to_cancel_list):
    # Create empty dicts
    inventory_requested_dict = {}
    inventory_over_min_dict = {}
    inventory_cancelled_dict = {}
    # Iterate over raw data sheet
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=(raw_data_sheet.max_row)):
        po = row[0].value
        model_number = row[2].value
        quantity_requested = row[9].value
        # Populate dict for 'Number of Units: Requested by Amazon'
        if model_number not in inventory_requested_dict.keys():
            inventory_requested_dict[model_number] = quantity_requested
        else:
            inventory_requested_dict[model_number] = (
                inventory_requested_dict[model_number] + quantity_requested
            )
        # Populate dict for 'Number of Unit: from POs > $380'
        if po not in orders_to_cancel_list:
            if model_number in inventory_over_min_dict.keys():
                inventory_over_min_dict[model_number] = (
                    inventory_over_min_dict[model_number] + quantity_requested
                )
            else:
                inventory_over_min_dict[model_number] = quantity_requested
        # Populate dict for inventory cancelled
        if po in orders_to_cancel_list:
            if model_number in inventory_cancelled_dict.keys():
                inventory_cancelled_dict[model_number] = (
                    inventory_cancelled_dict[model_number] + quantity_requested
                )
            else:
                inventory_cancelled_dict[model_number] = quantity_requested
    return (inventory_requested_dict, inventory_over_min_dict, inventory_cancelled_dict)


# Print the sorted 'Inventory Requested (from non-cancelled POs)' dictionary
def print_inventory_to_confirm(inventory_dicts, inv_heading_title):
    inventory_over_min_dict = inventory_dicts[1]
    print(inv_heading_title)
    longest_product_name = 0
    for key in inventory_over_min_dict.keys():
        if key is not None and len(key) > longest_product_name:
            longest_product_name = len(key)
    def custom_sort(item):
        key, value = item
        return (value is None, value)
    sorted_inventory_over_min_dict = dict(
        sorted(
            inventory_over_min_dict.items(), key=custom_sort, reverse=True
        )
    )

    for key, value in sorted_inventory_over_min_dict.items():
        if key is not None and value != 0:
            product = key
            spaces_needed = longest_product_name - len(product)
            extra_space = ""
            for space in range(spaces_needed):
                extra_space = extra_space + " "
            print(product, extra_space, " : ", value, sep="")


# Create a new sheet (in new_wb), with the POs and Requested Units
def create_inventory_sheet(new_wb):
    inventory_sheet = new_wb.create_sheet("Inv to Confirm")
    new_wb.active = new_wb["Inv to Confirm"]
    inventory_sheet.cell(row=1, column=1).value = "Model Number"
    inventory_sheet.cell(
        row=1, column=2
    ).value = "Number of Units: \n Requested by Amazon"
    inventory_sheet.cell(row=1, column=3).value = "Number of Units: \n from POs > $380"
    inventory_sheet.cell(row=1, column=4).value = "Number of Units: \n In Stock"
    inventory_sheet.cell(
        row=1, column=5
    ).value = "Number of Units: \n In Stock and from POs > $380"
    # Format the sheet nicely
    inventory_sheet.row_dimensions[1].height = 30
    for row in inventory_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
        for cell in row:
            x = cell.coordinate
            inventory_sheet[x].font = Font(bold=True)
            inventory_sheet.column_dimensions[x[0]].width = 30
            inventory_sheet["A1"].fill = PatternFill("solid", start_color="CCE5FF")
            inventory_sheet["B1"].fill = PatternFill("solid", start_color="CCE5FF")
            inventory_sheet["C1"].fill = PatternFill("solid", start_color="CCE5FF")
            inventory_sheet["D1"].fill = PatternFill("solid", start_color="FFCCE5")
            inventory_sheet["E1"].fill = PatternFill("solid", start_color="E8E8E8")
            for y in range(inventory_sheet.max_row + 1):
                for z in range(inventory_sheet.max_row + 1):
                    for a in range(1, 6):
                        inventory_sheet.cell(row=z + 1, column=a).alignment = Alignment(
                            horizontal="center", wrap_text=True, vertical="center"
                        )
    return inventory_sheet


# Add data to 'Inv to Confirm' sheet
def populate_inventory_sheet(inventory_sheet, inventory_dicts):
    inventory_requested_dict = inventory_dicts[0]
    inventory_over_min_dict = inventory_dicts[1]
    # inventory_cancelled_dict = inventory_dicts[2]
    # Sort inventory requested dict
    def custom_sort(item):
        key, value = item
        return (value is None, value)
    sorted_inventory_requested_dict = dict(
        sorted(
            inventory_requested_dict.items(), key=custom_sort, reverse=True
        )
    )
    # Populate cells
    for i in range(len(sorted_inventory_requested_dict)):
        # Add Model Number
        model_number = list(sorted_inventory_requested_dict.keys())[i]
        inventory_sheet.cell(row=i + 2, column=1).value = model_number
        # Add 'Number of Units: Requested by Amazon'
        inventory_sheet.cell(row=i + 2, column=2).value = inventory_requested_dict[
            model_number
        ]
        # Add 'Number of Units: from POs > min value' + 'Number of Units: In Stock"
        if model_number not in inventory_over_min_dict:
            inventory_sheet.cell(row=i + 2, column=3).value = 0
            inventory_sheet.cell(row=i + 2, column=4).value = 0
        else:
            inventory_sheet.cell(row=i + 2, column=3).value = inventory_over_min_dict[
                model_number
            ]
            inventory_sheet.cell(row=i + 2, column=4).value = inventory_over_min_dict[
                model_number
            ]
    # Format new values
    for row in inventory_sheet.iter_cols(
        min_row=2, max_row=inventory_sheet.max_row, min_col=4, max_col=4
    ):
        for cell in row:
            if cell.value is None:
                break
            elif cell.value == 0:
                cell.font = Font(color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="FFFDD0")
            else:
                cell.fill = PatternFill("solid", start_color="FFFDD0")
                cell.font = Font(color="A9A9A9")


# Create new sheet, containing the POs to Confirm
def create_pos_to_confirm_unadjusted_sheet(new_wb):
    unadjusted_pos_sheet = new_wb.create_sheet("POs to Confirm")
    unadjusted_pos_sheet.cell(row=1, column=1).value = "PO Number"
    unadjusted_pos_sheet.cell(row=1, column=2).value = "Accepted/Cancelled"
    unadjusted_pos_sheet.cell(row=1, column=3).value = "All Items Accepted?"
    unadjusted_pos_sheet.cell(row=1, column=4).value = "Model Number"
    unadjusted_pos_sheet.cell(row=1, column=5).value = "Quantity Requested"
    unadjusted_pos_sheet.cell(row=1, column=6).value = "Quantity Confirmed"
    unadjusted_pos_sheet.cell(row=1, column=7).value = "Currency"
    # Format the sheet nicely
    for row in unadjusted_pos_sheet.iter_cols(
        min_row=1, max_row=1, min_col=1, max_col=7
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            x = cell.coordinate
            unadjusted_pos_sheet[x].font = Font(bold=True)
            unadjusted_pos_sheet.column_dimensions[x[0]].width = 20
            cell.fill = PatternFill("solid", start_color="ade6d4")
    return unadjusted_pos_sheet


# Populate sheet 'POs to Confirm' with data
def populate_pos_to_confirm_unadjusted_sheet(
    raw_data_sheet, unadjusted_pos_sheet, orders_to_cancel_list
):
    # Iterate through values in raw data sheet to fill in base data
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=(raw_data_sheet.max_row)):
        po_number = row[0].value
        model_number = row[2].value
        requested_quantity = row[9].value
        row_number = row[0].row
        currency = row[31].value
        unadjusted_pos_sheet.cell(row=row_number, column=1).value = po_number
        unadjusted_pos_sheet.cell(row=row_number, column=3).font = Font(color="008000")
        unadjusted_pos_sheet.cell(row=row_number, column=4).value = model_number
        unadjusted_pos_sheet.cell(row=row_number, column=5).value = requested_quantity
        unadjusted_pos_sheet.cell(row=row_number, column=6).value = requested_quantity
        if po_number in orders_to_cancel_list:
            unadjusted_pos_sheet.cell(row=row_number, column=2).value = "CANCEL"
            unadjusted_pos_sheet.cell(row=row_number, column=2).font = Font(
                color="FF0000"
            )
            unadjusted_pos_sheet.cell(row=row_number, column=6).value = 0
        else:
            unadjusted_pos_sheet.cell(row=row_number, column=2).value = "ACCEPT"
            unadjusted_pos_sheet.cell(row=row_number, column=2).font = Font(
                color="008000"
            )
            unadjusted_pos_sheet.cell(row=row_number, column=3).value = "YES"
        if currency == "USD":
            unadjusted_pos_sheet.cell(row=row_number, column=7).value = "USD"
            unadjusted_pos_sheet.cell(row=row_number, column=7).font = Font(
                color="0000FF"
            )
        else:
            unadjusted_pos_sheet.cell(row=row_number, column=7).value = "CAD"
            unadjusted_pos_sheet.cell(row=row_number, column=7).font = Font(
                color="FF0000"
            )
    # Format 'POs to Confirm Sheet' -> alternate background color of POs in 'POs to Confirm'
    prev = ""
    curr_color = "FFFFFF"
    alt_color = "E8E8E8"
    saved_color = "FFFFFF"
    for row in unadjusted_pos_sheet.iter_rows(
        min_row=2, max_row=(raw_data_sheet.max_row)
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            po_number = row[0].value
            if prev == "":
                cell.fill = PatternFill("solid", start_color=curr_color)
            elif po_number == prev:
                cell.fill = PatternFill("solid", start_color=curr_color)
            else:
                cell.fill = PatternFill("solid", start_color=alt_color)
                saved_color = curr_color
                curr_color = alt_color
                alt_color = saved_color
            prev = po_number


# Update the columns 'Quantity Confirmed' + 'Quantity Cancelled' in the 'PO Raw Data' sheet
def update_raw_data_sheet(raw_data_sheet, orders_to_cancel_list):
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=(raw_data_sheet.max_row)):
        for cell in row:
            po_number = row[0].value
            quantity_ordered = row[9].value
            if po_number in orders_to_cancel_list:
                # Make the 'Quantity Ordered' column value 0
                row[11].value = 0
                # Make the 'Quantity Canceled' column value equal to the 'Quantity Ordered'
                row[12].value = quantity_ordered


# Save the workbook
def save_new_wb(new_wb, new_wb_filename):
    new_wb.save(new_wb_filename)
