from openpyxl import load_workbook


# Open .xlsx file 
wb = load_workbook(filename = 'PurchaseOrderItems.xlsx')


# Open first sheet of .xlsx file
sheet = wb.active


# Tests
assert sheet['G1'].value == 'Model Number',  'Cell G1 should be Model Number'
assert sheet['P1'].value == 'Accepted Quantity', 'Cell P1 should be Accepted Quantity'


# Create empty dictionary
inventory_dict = {}


# Limit row iteration to 1000 rows
if sheet.max_row > 1000:
    row_limit = 1000
else:
    row_limit = sheet.max_row 


# Iterate through values in column G and P (i.e. Model Number and Accepted Quantity)
for row in sheet.iter_rows(min_row=2, max_row=(row_limit)):
    product = row[6].value
    quantity = row[15].value
    # if product in dictionary, update quantity
    if product in inventory_dict.keys():
        inventory_dict[product] = inventory_dict[product] + quantity
    # if product not in dictionary, add it to dictionary
    else:
        inventory_dict[product] = quantity


# Print the dictionary
print('Amazon Orders\n-------------')
for key, value in inventory_dict.items():
        print(key, ':', value)


