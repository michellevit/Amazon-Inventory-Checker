from multiprocessing.sharedctypes import Value
from re import A
from openpyxl import load_workbook
import os
from datetime import date
from amazon_script_1 import *
import operator


def main():
    # Check if 'POs to Confirm [DATE].xlsx' file exists (needed to run this program)
    today = date.today()
    pos_to_confirm_filename = (
        "POs to Confirm - " + today.strftime("%B %d, %Y") + ".xlsx"
    )
    # Check if the file exists
    check_file(pos_to_confirm_filename)
    # Get workbook + sheets from within the workbook
    new_wb = load_workbook(filename=pos_to_confirm_filename)
    raw_data_sheet = new_wb["PO Raw Data"]
    inv_to_confirm_sheet = new_wb["Inv to Confirm"]
    pos_to_confirm_sheet = new_wb["POs to Confirm"]
    # Create a dict for the units to cancel (from non-cancelled orders / orders over min threshold)
    units_to_cancel_dict = get_units_to_cancel(inv_to_confirm_sheet)
    # Improve formatting/readability of new workbook
    reformat_inv_to_confirm_sheet(inv_to_confirm_sheet)
    # Get PO values
    po_value_dict = get_po_values(raw_data_sheet)
    # Remove out of stock units - if any (from non-cancelled orders / orders over min threshold)
    remove_out_of_stock_units(raw_data_sheet, units_to_cancel_dict, po_value_dict)
    # Recalculate orders over min threshold
    sorted_po_list = get_and_sort_po_values(raw_data_sheet)
    orders_to_cancel_list = create_list_pos_to_cancel(sorted_po_list)
    # Remove cancelled orders from PO Raw Data sheet
    update_raw_data_sheet(raw_data_sheet, orders_to_cancel_list)
    # Print data to terminal
    print_pos_to_confirm_final_version(raw_data_sheet)
    print_inventory_to_confirm_final_version(raw_data_sheet)
    # Update newly created sheets (allows for manual entry, if the user wishes, and makes it easy to verify any issues)
    update_inventory_to_confirm_sheet(raw_data_sheet, inv_to_confirm_sheet)
    update_pos_to_confirm_sheet(raw_data_sheet, pos_to_confirm_sheet)
    save_new_wb(new_wb, pos_to_confirm_filename)
    # Update Vendor Download Sheets with correct values
    po_dict = create_po_dict(raw_data_sheet)
    # Update Vendor Download Sheets with correct values
    usd_file, cad_file = get_vendor_download_file()
    if usd_file != "empty":
        update_vendor_download_sheet(usd_file, po_dict)
    if cad_file != "empty":
        update_vendor_download_sheet(cad_file, po_dict)


def check_file(pos_to_confirm_filename):
    if not os.path.exists(pos_to_confirm_filename):
        raise Exception(
            "Error: The required file does not exist - please make sure you have run the first script properly before running this one."
        )


def get_units_to_cancel(inv_to_confirm_sheet):
    units_to_cancel_dict = {}
    # Iterate through data from 'Inv to Confirm' Sheet:
    for row in inv_to_confirm_sheet.iter_rows(
        min_row=2, max_row=inv_to_confirm_sheet.max_row
    ):
        if row[0].value == None:
            break
        else:
            model_number = row[0].value
            units_from_pos_under_380 = row[2].value
            units_in_stock = row[3].value
            if units_in_stock < units_from_pos_under_380:
                units_to_cancel = units_from_pos_under_380 - units_in_stock
                if model_number in units_to_cancel_dict.keys():
                    units_to_cancel_dict[model_number] = (
                        units_to_cancel_dict[model_number] + units_to_cancel
                    )
                else:
                    units_to_cancel_dict[model_number] = units_to_cancel
    return units_to_cancel_dict


def reformat_inv_to_confirm_sheet(inv_to_confirm_sheet):
    inv_to_confirm_sheet["C1"].fill = PatternFill("solid", start_color="CCE5FF")
    inv_to_confirm_sheet["D1"].fill = PatternFill("solid", start_color="CCE5FF")
    inv_to_confirm_sheet["E1"].fill = PatternFill("solid", start_color="C1E1C1")
    # Remove background color from cells in column D
    for row in inv_to_confirm_sheet.iter_cols(
        min_row=2, max_row=inv_to_confirm_sheet.max_row, min_col=4, max_col=4
    ):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(color="000000")


def get_po_values(raw_data_sheet):
    po_value_dict = {}
    for row in raw_data_sheet.iter_rows(min_row=2, max_row=(raw_data_sheet.max_row)):
        po_number = row[0].value
        cost = row[8].value
        quantity_confirmed = row[11].value
        if po_number in po_value_dict.keys():
            po_value_dict[po_number] = po_value_dict[po_number] + (
                cost * quantity_confirmed
            )
        else:
            po_value_dict[po_number] = cost * quantity_confirmed
    return po_value_dict


def remove_out_of_stock_units(raw_data_sheet, units_to_cancel_dict, po_value_dict):
    rows = list(raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row))
    rows = reversed(rows)
    for row in rows:
        model_number = row[2].value
        if model_number in units_to_cancel_dict.keys():
            po_number = row[0].value
            cost = row[8].value
            po_value = po_value_dict[po_number]
            units_to_cancel = units_to_cancel_dict[model_number]
            for x in range(units_to_cancel):
                if (po_value - cost) > min_po_value:
                    if row[11].value > 0:
                        po_value_dict[po_number] = po_value_dict[po_number] - cost
                        po_value = po_value_dict[po_number]
                        row[11].value = (
                            row[11].value - 1
                        )  # row[11] = 'Quantity Confirmed'
                        row[12].value = (
                            row[12].value + 1
                        )  # row[12] = 'Quantity Canceled'
                        units_to_cancel_dict[model_number] = (
                            units_to_cancel_dict[model_number] - 1
                        )
                    if units_to_cancel_dict[model_number] == 0:
                        del units_to_cancel_dict[model_number]
    # If there are remaining units to cancel, remove them starting from the PO with the lowest value
    if len(units_to_cancel_dict) != 0:
        low_to_high_po_value_list = sorted(po_value_dict.items(), key=lambda x: x[1])
        for index, tuple in enumerate(low_to_high_po_value_list):
            list_po_number = tuple[0]
            rows = list(
                raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row)
            )
            rows = reversed(rows)
            for row in rows:
                po_number = row[0].value
                model_number = row[2].value
                cost = row[8].value
                po_value = po_value_dict[po_number]
                if po_number == list_po_number:
                    if model_number in units_to_cancel_dict.keys():
                        units_to_cancel = units_to_cancel_dict[model_number]
                        for y in range(units_to_cancel):
                            if row[11].value > 0:
                                po_value_dict[po_number] = (
                                    po_value_dict[po_number] - cost
                                )
                                row[11].value = (
                                    row[11].value - 1
                                )  # row[11] = 'Quantity Ordered'
                                row[12].value = (
                                    row[12].value + 1
                                )  # row[12] = 'Quantity Canceled'
                                units_to_cancel_dict[model_number] = (
                                    units_to_cancel_dict[model_number] - 1
                                )
                                if units_to_cancel_dict[model_number] == 0:
                                    del units_to_cancel_dict[model_number]


def print_pos_to_confirm_final_version(raw_data_sheet):
    po_value_dict = get_po_values(raw_data_sheet)
    sorted_po_list = sorted(po_value_dict.items(), key=lambda x: x[1], reverse=True)
    cad_po_list = get_cad_pos(raw_data_sheet)
    orders_to_cancel_list = create_list_pos_to_cancel(sorted_po_list)
    po_heading_title = (
        "---------------\nAmazon: POs to Confirm/Cancel (confirmed)\n---------------"
    )
    print_pos_to_confirm(
        sorted_po_list, cad_po_list, orders_to_cancel_list, po_heading_title
    )


def print_inventory_to_confirm_final_version(raw_data_sheet):
    po_value_dict = get_po_values(raw_data_sheet)
    sorted_po_list = sorted(po_value_dict.items(), key=lambda x: x[1], reverse=True)
    orders_to_cancel_list = create_list_pos_to_cancel(sorted_po_list)
    inventory_dicts = create_inventory_tracker_dicts(
        raw_data_sheet, orders_to_cancel_list
    )
    inv_heading_title = "---------------\nAmazon: Inventory Requested (from non-cancelled POs - confirmed)\n---------------"
    print_inventory_to_confirm(inventory_dicts, inv_heading_title)


def update_inventory_to_confirm_sheet(raw_data_sheet, inv_to_confirm_sheet):
    confirmed_inventory_dict = {}
    for row in raw_data_sheet.iter_rows(
        min_row=2, max_row=(inv_to_confirm_sheet.max_row)
    ):
        model_number = row[2].value
        quantity_confirmed = row[11].value
        if model_number == None:
            break
        elif model_number in confirmed_inventory_dict.keys():
            confirmed_inventory_dict[model_number] = (
                confirmed_inventory_dict[model_number] + quantity_confirmed
            )
        else:
            confirmed_inventory_dict[model_number] = quantity_confirmed
    for row in inv_to_confirm_sheet.iter_rows(
        min_row=2, max_row=(inv_to_confirm_sheet.max_row)
    ):
        model_number = row[0].value
        model_number = row[0].value
        if model_number == None:
            break
        else:
            row[4].value = confirmed_inventory_dict[model_number]
            row[4].fill = PatternFill("solid", start_color="FFFDD0")


def update_pos_to_confirm_sheet(raw_data_sheet, pos_to_confirm_sheet):
    sorted_po_list = get_and_sort_po_values(raw_data_sheet)
    orders_to_cancel_list = create_list_pos_to_cancel(sorted_po_list)
    rows = list(
        pos_to_confirm_sheet.iter_rows(min_row=2, max_row=pos_to_confirm_sheet.max_row)
    )
    for row in rows:
        new_po_number = row[0].value
        new_model_number = row[3].value
        if new_po_number in orders_to_cancel_list:
            row[1].value = "CANCELLED"
            row[1].font = Font(color="FF0000")
            row[2].value = " "
            row[5].value = 0
        else:
            crows = list(
                raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row)
            )
            for crow in crows:
                old_po_number = crow[0].value
                old_model_number = crow[2].value
                old_quantity_confirmed = crow[11].value
                if (new_po_number == old_po_number) and (
                    new_model_number == old_model_number
                ):
                    row[5].value = old_quantity_confirmed
                    new_quantity_requested = row[4].value
                    new_quantity_confirmed = row[5].value
                    if new_quantity_requested != new_quantity_confirmed:
                        row[5].fill = PatternFill("solid", start_color="FFFDD0")
                        row[2].value = "NO"
                        row[2].font = Font(color="FF0000")





# Create dict of pos with dict inside of po line items
def create_po_dict(raw_data_sheet):
    rows = list(raw_data_sheet.iter_rows(min_row=2, max_row=raw_data_sheet.max_row + 1))
    po_dict = {}
    line_items = {}
    previous_po = ""
    x = 1
    for row in rows:
        if row[0].value != None:
            current_po = row[0].value
            if (previous_po == current_po) or (previous_po == ""):
                model_number = row[2].value
                quantity_confirmed = row[11].value
                line_items[model_number] = quantity_confirmed
                previous_po = current_po
            else:
                po_dict[previous_po] = line_items
                line_items = {}
                model_number = row[2].value
                quantity_confirmed = row[11].value
                line_items[model_number] = quantity_confirmed
                previous_po = current_po
        else:
            po_dict[previous_po] = line_items
            break
    return po_dict


def get_vendor_download_file():
    my_dir = os.listdir()
    amazon_files = []
    usd_file = "empty"
    cad_file = "empty"
    for x in my_dir:
        if "VendorDownload" in x:
            amazon_file = x
            amazon_files.append(amazon_file)
    if len(amazon_files) >= 1:
        amazon_file_1 = amazon_files[0]
        wb_1 = load_workbook(filename=amazon_file_1)
        ws_1 = wb_1.active
        currency = ws_1["AF4"].value
        if currency == "USD":
            usd_file = amazon_files[0]
        elif currency == "CAD":
            cad_file = amazon_files[0]
    if len(amazon_files) == 2:
        amazon_file_2 = amazon_files[1]
        wb_2 = load_workbook(filename=amazon_file_2)
        ws_2 = wb_2.active
        currency = ws_2["AF4"].value
        if currency == "USD":
            if usd_file != "empty":
                raise Exception("Error: There are 2 USD Vendor Download Files")
            usd_file = amazon_files[1]
        elif currency == "CAD":
            if cad_file != "empty":
                raise Exception("Error: There are 2 CAD Vendor Download Files")
            cad_file = amazon_files[1]
    return usd_file, cad_file


def update_vendor_download_sheet(file, po_dict):
    wb = load_workbook(filename=file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row))
    for row in rows:
        # VD Column A (Column #0)= PO Number
        # VD Column C (Column #2) = Model Number
        # VD Column L (Column #11) = Quantity Confirmed
        # VD Column P (Column #15) = Hand Off End
        # VD Column R (Column #17) = Expected Hand Off End
        # VD Column S (Column #18) = Availability Status (AC/OS)
        # VD Column X (Column #23) = Condition (confirmed/unconfirmed)
        po = row[0].value
        model_number = row[2].value
        # Get PO line items
        line_items = po_dict[po]
        quantity_confirmed = line_items[model_number]
        # Update VD sheet w/ actual quantity confirmed
        row[11].value = quantity_confirmed
        # Update VD sheet "Expected Hand Off Date" with the 'Hand Off End' date
        row[17].value = row[15].value
        # Update 'Availability Status' if zero quantity is accepted
        if quantity_confirmed == 0:
            row[18].value = "OS - Cancelled: Out of stock"
        # Update 'Condition' to "Confirmed"
        row[23].value = "Confirmed"
        wb.save(filename=file
                )



# Save the workbook
def save_new_wb(new_wb, pos_to_confirm_filename):
    new_wb.save(pos_to_confirm_filename)



main()
